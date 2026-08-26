from __future__ import annotations

import argparse
import bisect
import gc
import hashlib
import io
import json
import re
import shutil
import subprocess
import tempfile
import unicodedata
import warnings
from collections import defaultdict
from pathlib import Path

import openpyxl
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
DATA_PATH = ROOT / "public/parts-data.json"
FIGURES_DIR = ROOT / "public/figures"
DRY_RUN = False

PART_CODE_HEADERS = (
    "物料编码",
    "零件代号",
    "备件编码",
    "partnumber",
    "partno",
    "materialcode",
    "sparepartcode",
    "itemno",
)
POSITION_HEADERS = ("序号", "no", "serialnumber", "s/n", "sn")
VIN_RE = re.compile(r"[A-HJ-NPR-Z0-9]{17}")
CODE_TOKEN_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._/\-]{2,79}$")


def text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\u200b", "").replace("\ufeff", "").strip()


def compact(value: object) -> str:
    return re.sub(r"[\s:：.]+", "", text(value)).lower()


def clean_code(value: object) -> str:
    return re.sub(r"\s+", "", text(value))


def unique(values: list[str]) -> list[str]:
    return list(dict.fromkeys(value for value in values if value))


def is_header(row: list[str]) -> bool:
    if len(row) < 2:
        return False
    first = compact(row[0])
    second = compact(row[1])
    return any(token in first for token in POSITION_HEADERS) and any(
        token in second for token in PART_CODE_HEADERS
    )


def looks_like_code(value: str) -> bool:
    value = clean_code(value)
    if not value or len(value) > 80 or not CODE_TOKEN_RE.fullmatch(value):
        return False
    if not re.search(r"\d", value):
        return False
    return True


def is_part_number(value: str) -> bool:
    value = clean_code(value)
    if not looks_like_code(value):
        return False
    lowered = value.lower()
    if any(token in lowered for token in ("partnumber", "partno", "materialcode", "itemno")):
        return False
    if any("\u4e00" <= char <= "\u9fff" for char in value):
        return False
    if value.isdigit() and len(value) < 5:
        return False
    return True


CATALOG_YEAR_RE = re.compile(r"_(20\d{2})")
CATALOG_MODEL_RE = re.compile(r"\bLCK[A-Z0-9]+\b", re.IGNORECASE)
CATALOG_ENGINE_RE = re.compile(r"[-\s]([A-Z]\d{1,2})\s*engine", re.IGNORECASE)
CATALOG_BODY_RE = re.compile(r"catalogs?\s+for\s+(?P<body>.+?)(?:_(?=20\d{2})|$)", re.IGNORECASE)


def derive_catalog_metadata(catalog: str) -> tuple[dict, list[str]]:
    """Propose catalog metadata from the workbook name.

    Supplier files are named like "Parts catalog for LCK6128H coach-L9 engine_2022_52 units".
    Only rules that hold across the existing catalogs are applied; every field that
    could not be inferred is returned in `uncertain` so the caller can flag it for
    review instead of writing a plausible-looking guess.
    """
    model_match = CATALOG_MODEL_RE.search(catalog)
    year_match = CATALOG_YEAR_RE.search(catalog)
    engine_match = CATALOG_ENGINE_RE.search(catalog)
    model = model_match.group(0).upper() if model_match else ""
    year = year_match.group(1) if year_match else ""

    # Battery-electric models carry EV in the designation and have no engine code.
    if model and re.search(r"EV", model):
        engine = "Electric"
    else:
        engine = engine_match.group(1).upper() if engine_match else ""

    body_match = CATALOG_BODY_RE.search(catalog)
    body = text(body_match.group("body")) if body_match else ""
    if model:
        body = re.sub(re.escape(model), "", body, flags=re.IGNORECASE).strip(" -_")
    body = re.sub(r"[-\s]*\b[A-Z]?\d*\s*engine\b", "", body, flags=re.IGNORECASE)
    body = re.sub(r"_[A-Z0-9-]{6,}(?:（[^）]*）)?\s*$", "", body, flags=re.IGNORECASE)
    body = re.sub(r"\s+", " ", body).strip(" -_")
    # "intercity" and "coach" are bare body styles; "city bus", "Ebus", "E-chassis"
    # and "mini bus" already name the vehicle.
    if (body and "-" not in body and " " not in body
            and re.fullmatch(r"[A-Za-z]+", body) and not re.search(r"bus", body, re.IGNORECASE)):
        body = f"{body} bus"
    vehicle_type = " ".join(part for part in (model, body) if part).strip()

    uncertain = [name for name, value in (("model", model), ("year", year)) if not value]
    if model and not body:
        uncertain.append("vehicleType")
    metadata = {
        "model": model,
        "year": year,
        "engine": engine,
        "vehicleType": vehicle_type or model,
        "catalog": catalog,
        "vinNumbers": [],
    }
    return metadata, uncertain


def is_assembly_row(row: list[str]) -> bool:
    if not row or is_header(row):
        return False
    first = text(row[0])
    second = text(row[1]) if len(row) > 1 else ""
    third = text(row[2]) if len(row) > 2 else ""
    fourth = text(row[3]) if len(row) > 3 else ""
    if not first or second:
        return False
    if third or fourth:
        return looks_like_code(first.split()[0])
    pieces = first.split(maxsplit=1)
    return len(pieces) == 2 and looks_like_code(pieces[0]) and bool(pieces[1].strip())


def split_assembly(row: list[str]) -> tuple[str, str, str]:
    first = text(row[0])
    chinese = text(row[2]) if len(row) > 2 else ""
    english = text(row[3]) if len(row) > 3 else ""
    pieces = first.split(maxsplit=1)
    if pieces and looks_like_code(pieces[0]) and (not chinese and not english):
        code = clean_code(pieces[0])
        title = pieces[1].strip() if len(pieces) > 1 else ""
        return code, title, title
    code = clean_code(pieces[0] if pieces and looks_like_code(pieces[0]) else first)
    title = english or chinese or (pieces[1].strip() if len(pieces) > 1 else first)
    chinese_title = chinese or title
    return code, title, chinese_title


def normalize_position(value: str) -> str:
    value = text(value)
    if re.fullmatch(r"\d+\.0", value):
        return value[:-2]
    return value


def find_detail_sheet(workbook: openpyxl.Workbook):
    for sheet in workbook.worksheets:
        if "detail" in sheet.title.lower():
            return sheet
    for sheet in workbook.worksheets:
        if "summary" in sheet.title.lower() or "汇总" in sheet.title:
            return sheet
    return max(workbook.worksheets, key=lambda sheet: sheet.max_row)


def rows_from_sheet(sheet) -> list[list[str]]:
    return [
        [text(value) for value in row]
        for row in sheet.iter_rows(min_col=1, max_col=min(10, sheet.max_column), values_only=True)
    ]


def primary_headers(rows: list[list[str]]) -> list[int]:
    headers = [index for index, row in enumerate(rows) if is_header(row)]
    primary: list[int] = []
    for index in headers:
        if not primary or index - primary[-1] > 1:
            primary.append(index)
    return primary


def preceding_assembly(rows: list[list[str]], header_index: int) -> int | None:
    lower_bound = max(0, header_index - 160)
    for index in range(header_index - 1, lower_bound - 1, -1):
        if is_assembly_row(rows[index]):
            return index
        if is_header(rows[index]):
            break
    return None


def extract_vins(workbook: openpyxl.Workbook, fallback: list[str]) -> list[str]:
    values: list[str] = []
    for sheet in workbook.worksheets:
        if "instruction" not in sheet.title.lower():
            continue
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                for match in VIN_RE.findall(text(value).upper()):
                    values.append(match)
    # Keep the VIN coverage already verified in the live catalog and merge any
    # newly listed chassis numbers from the update workbook. Some supplier
    # instruction sheets only show a representative VIN even when the catalog
    # applies to a full fleet.
    return unique([*fallback, *values])


def figure_candidates(sheet, header_indexes: list[int]) -> dict[int, tuple[int, bytes]]:
    selected: dict[int, tuple[int, bytes]] = {}
    for image in getattr(sheet, "_images", []):
        anchor = getattr(getattr(image, "anchor", None), "_from", None)
        if anchor is None:
            continue
        anchor_index = int(anchor.row)
        position = bisect.bisect_right(header_indexes, anchor_index)
        if position >= len(header_indexes):
            continue
        header_index = header_indexes[position]
        if header_index - anchor_index > 180:
            continue
        try:
            payload = image._data()
            with Image.open(io.BytesIO(payload)) as source:
                area = source.width * source.height
        except Exception:
            continue
        previous = selected.get(header_index)
        if previous is None or area > previous[0]:
            selected[header_index] = (area, payload)
    return selected


def save_figure(payload: bytes) -> str:
    with Image.open(io.BytesIO(payload)) as source:
        source.load()
        if source.mode not in ("RGB", "RGBA"):
            source = source.convert("RGBA" if "transparency" in source.info else "RGB")
        if source.width > 1800:
            height = max(1, round(source.height * 1800 / source.width))
            source.thumbnail((1800, height), Image.Resampling.LANCZOS)
        output = io.BytesIO()
        source.save(output, format="WEBP", quality=88, method=6)
    payload = output.getvalue()
    name = f"{hashlib.sha256(payload).hexdigest()[:16]}.webp"
    target = FIGURES_DIR / name
    if not target.exists() and not DRY_RUN:
        target.write_bytes(payload)
    return f"/figures/{name}"


def old_figure_index(old_data: dict) -> dict[tuple[str, str], str]:
    figures: dict[tuple[str, str], str] = {}
    for group in old_data.get("groups", {}).values():
        figure = text(group.get("figure"))
        code = clean_code(group.get("code"))
        catalog = text(group.get("catalog"))
        if figure and code:
            figures[(catalog, code)] = figure
    return figures


def occurrence(
    *,
    part_number: str,
    description: str,
    chinese: str,
    quantity: str,
    unit: str,
    notes: str,
    assembly: str,
    assembly_code: str,
    metadata: dict,
    group_id: str,
    position: str,
) -> dict:
    vins = metadata.get("vinNumbers", [])
    return {
        "partNumber": part_number,
        "description": description or chinese or part_number,
        "descriptionChinese": chinese or description or part_number,
        "quantity": quantity,
        "unit": unit,
        "notes": notes,
        "assembly": assembly,
        "assemblyCode": assembly_code,
        "model": metadata.get("model", ""),
        "year": metadata.get("year", ""),
        "engine": metadata.get("engine", ""),
        "vehicleType": metadata.get("vehicleType", ""),
        "representativeVin": vins[0] if vins else "",
        "vinCount": len(vins),
        "catalog": metadata["catalog"],
        "groupId": group_id,
        "position": position,
    }


def parse_workbook(
    path: Path,
    metadata: dict,
    start_group: int,
    fallback_figures: dict[tuple[str, str], str],
) -> tuple[list[dict], dict[str, dict], int, list[str]]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = find_detail_sheet(workbook)
    rows = rows_from_sheet(sheet)
    headers = primary_headers(rows)
    metadata = dict(metadata)
    metadata["vinNumbers"] = extract_vins(workbook, metadata.get("vinNumbers", []))
    image_map = figure_candidates(sheet, headers)
    occurrences: list[dict] = []
    groups: dict[str, dict] = {}
    group_number = start_group

    for header_position, header_index in enumerate(headers):
        assembly_index = preceding_assembly(rows, header_index)
        assembly_row = rows[assembly_index] if assembly_index is not None else [""] * 7
        assembly_code, assembly_title, assembly_chinese = split_assembly(assembly_row)
        group_id = f"g{group_number}"
        group_number += 1

        figure = ""
        candidate = image_map.get(header_index)
        if candidate:
            try:
                figure = save_figure(candidate[1])
            except Exception:
                figure = ""
        if not figure:
            figure = fallback_figures.get((metadata["catalog"], assembly_code), "")

        part_numbers: list[str] = []
        if is_part_number(assembly_code) and not assembly_code.upper().startswith("LCK"):
            part_numbers.append(assembly_code)
            occurrences.append(
                occurrence(
                    part_number=assembly_code,
                    description=assembly_title,
                    chinese=assembly_chinese,
                    quantity="",
                    unit="",
                    notes="",
                    assembly=assembly_title,
                    assembly_code=assembly_code,
                    metadata=metadata,
                    group_id=group_id,
                    position="",
                )
            )

        data_start = header_index + 1
        while data_start < len(rows) and is_header(rows[data_start]):
            data_start += 1
        next_header = headers[header_position + 1] if header_position + 1 < len(headers) else len(rows)
        data_end = next_header
        for index in range(data_start, next_header):
            if is_assembly_row(rows[index]):
                data_end = index
                break

        for row in rows[data_start:data_end]:
            if len(row) < 2:
                continue
            part_number = clean_code(row[1])
            if not is_part_number(part_number):
                continue
            position_value = normalize_position(row[0])
            chinese = text(row[2]) if len(row) > 2 else ""
            english = text(row[3]) if len(row) > 3 else ""
            quantity = text(row[4]) if len(row) > 4 else ""
            unit = text(row[5]) if len(row) > 5 else ""
            notes = text(row[6]) if len(row) > 6 else ""
            part_numbers.append(part_number)
            occurrences.append(
                occurrence(
                    part_number=part_number,
                    description=english or chinese,
                    chinese=chinese or english,
                    quantity=quantity,
                    unit=unit,
                    notes=notes,
                    assembly=assembly_title,
                    assembly_code=assembly_code,
                    metadata=metadata,
                    group_id=group_id,
                    position=position_value,
                )
            )

        groups[group_id] = {
            "catalog": metadata["catalog"],
            "code": assembly_code,
            "title": assembly_title,
            "parts": sorted(set(part_numbers)),
            "figure": figure,
        }

    workbook.close()
    gc.collect()
    return occurrences, groups, group_number, metadata["vinNumbers"]


def retained_catalog(
    old_data: dict,
    catalog: str,
    start_group: int,
) -> tuple[list[dict], dict[str, dict], int]:
    old_groups = {
        group_id: group
        for group_id, group in old_data.get("groups", {}).items()
        if group.get("catalog") == catalog
    }
    group_map: dict[str, str] = {}
    groups: dict[str, dict] = {}
    group_number = start_group
    for old_id, group in old_groups.items():
        new_id = f"g{group_number}"
        group_number += 1
        group_map[old_id] = new_id
        groups[new_id] = {**group}
    occurrences: list[dict] = []
    for part in old_data.get("parts", []):
        for item in part.get("occurrences", []):
            if item.get("catalog") != catalog:
                continue
            copied = {**item, "groupId": group_map.get(item.get("groupId", ""), "")}
            occurrences.append(copied)
    return occurrences, groups, group_number


def repair_blank_groups(groups: dict[str, dict], occurrences: list[dict], old_data: dict) -> None:
    """Restore section metadata omitted by some supplier workbook exports.

    A few Excel drawings contain the assembly caption as a shape rather than a
    cell value. In that case openpyxl can read the parts and figure but not the
    caption. Match the section to the previous catalog by overlapping parts;
    for a genuinely new section, inherit the nearest preceding assembly in the
    same catalog. Empty header-only sections are removed.
    """
    old_by_catalog: dict[str, list[tuple[set[str], dict]]] = defaultdict(list)
    for old_group in old_data.get("groups", {}).values():
        old_by_catalog[text(old_group.get("catalog"))].append(
            (set(old_group.get("parts", [])), old_group)
        )

    last_named: dict[str, tuple[str, str]] = {}
    removed: set[str] = set()
    for group_id, group in groups.items():
        catalog = text(group.get("catalog"))
        part_numbers = set(group.get("parts", []))
        if not part_numbers:
            removed.add(group_id)
            continue
        if group.get("code") or group.get("title"):
            last_named[catalog] = (text(group.get("code")), text(group.get("title")))
            continue

        best_overlap = 0
        best_group: dict | None = None
        for old_parts, old_group in old_by_catalog.get(catalog, []):
            overlap = len(part_numbers & old_parts)
            if overlap > best_overlap:
                best_overlap = overlap
                best_group = old_group
        if best_group is not None:
            group["code"] = text(best_group.get("code"))
            group["title"] = text(best_group.get("title"))
        elif catalog in last_named:
            group["code"], group["title"] = last_named[catalog]
        if group.get("code") or group.get("title"):
            last_named[catalog] = (text(group.get("code")), text(group.get("title")))

    for group_id in removed:
        groups.pop(group_id, None)

    for item in occurrences:
        group = groups.get(item.get("groupId", ""))
        if group is None:
            continue
        item["assemblyCode"] = text(group.get("code"))
        item["assembly"] = text(group.get("title"))


def aggregate_parts(occurrences: list[dict]) -> list[dict]:
    by_number: dict[str, list[dict]] = defaultdict(list)
    for item in occurrences:
        by_number[item["partNumber"]].append(item)

    parts: list[dict] = []
    for part_number in sorted(by_number, key=lambda value: value.casefold()):
        items = by_number[part_number]
        english = next(
            (
                text(item["description"])
                for item in items
                if re.search(r"[A-Za-z]", text(item["description"]))
            ),
            text(items[0]["description"]),
        )
        chinese = next(
            (text(item["descriptionChinese"]) for item in items if text(item["descriptionChinese"])),
            english,
        )
        parts.append(
            {
                "partNumber": part_number,
                "description": english or chinese or part_number,
                "descriptionChinese": chinese or english or part_number,
                "models": unique([text(item["model"]) for item in items]),
                "years": unique([text(item["year"]) for item in items]),
                "assemblies": unique([text(item["assembly"]) for item in items]),
                "occurrences": items,
            }
        )
    return parts


def convert_legacy_files(source_dir: Path, destination: Path) -> list[Path]:
    converted: list[Path] = []
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    for source in source_dir.rglob("*.xls"):
        if not soffice:
            raise RuntimeError(f"LibreOffice is required to convert {source.name}")
        profile = destination / f"profile-{hashlib.sha1(source.name.encode()).hexdigest()[:8]}"
        profile.mkdir(parents=True, exist_ok=True)
        subprocess.run(
            [
                soffice,
                f"-env:UserInstallation=file://{profile}",
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(destination),
                str(source),
            ],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
        )
        target = destination / f"{source.stem}.xlsx"
        if not target.exists():
            raise RuntimeError(f"Conversion did not create {target.name}")
        converted.append(target)
    return converted


def main() -> None:
    parser = argparse.ArgumentParser(description="Import updated Zhongtong parts catalogs")
    parser.add_argument("--source-dir", type=Path, required=True)
    parser.add_argument("--converted-dir", type=Path)
    parser.add_argument("--update-date", default="2026-08-28")
    parser.add_argument("--output", type=Path, default=DATA_PATH)
    parser.add_argument(
        "--meta",
        type=Path,
        help="JSON file mapping catalog name to metadata overrides for new catalogs",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Report what would change without writing parts-data.json or figures",
    )
    args = parser.parse_args()

    overrides: dict[str, dict] = {}
    if args.meta:
        overrides = json.loads(args.meta.read_text(encoding="utf-8"))

    global DRY_RUN
    DRY_RUN = args.dry_run
    old_data = json.loads(DATA_PATH.read_text(encoding="utf-8"))
    fallback_figures = old_figure_index(old_data)
    old_part_counts = {item["catalog"]: item.get("parts", 0) for item in old_data["catalogs"]}
    if not DRY_RUN:
        FIGURES_DIR.mkdir(parents=True, exist_ok=True)

    temp_context = None
    if args.converted_dir:
        converted_dir = args.converted_dir
    else:
        temp_context = tempfile.TemporaryDirectory(prefix="zhongtong-converted-")
        converted_dir = Path(temp_context.name)
        convert_legacy_files(args.source_dir, converted_dir)

    workbook_paths = {path.stem: path for path in args.source_dir.rglob("*.xlsx")}
    workbook_paths.update({path.stem: path for path in converted_dir.glob("*.xlsx")})

    all_occurrences: list[dict] = []
    all_groups: dict[str, dict] = {}
    catalogs: list[dict] = []
    group_number = 1
    updated: list[str] = []
    retained: list[str] = []
    added: list[dict] = []

    for previous_metadata in old_data["catalogs"]:
        catalog = previous_metadata["catalog"]
        metadata = {**previous_metadata}
        path = workbook_paths.get(catalog)
        if path:
            occurrences, groups, group_number, vin_numbers = parse_workbook(
                path, metadata, group_number, fallback_figures
            )
            for item in occurrences:
                item["vinCount"] = len(vin_numbers)
                item["representativeVin"] = vin_numbers[0] if vin_numbers else ""
            updated.append(catalog)
        else:
            occurrences, groups, group_number = retained_catalog(old_data, catalog, group_number)
            vin_numbers = metadata.get("vinNumbers", [])
            retained.append(catalog)

        all_occurrences.extend(occurrences)
        all_groups.update(groups)
        catalogs.append(
            {
                "model": metadata.get("model", ""),
                "year": metadata.get("year", ""),
                "engine": metadata.get("engine", ""),
                "vehicleType": metadata.get("vehicleType", ""),
                "catalog": catalog,
                "parts": len(occurrences),
                "vins": len(vin_numbers),
                "vinNumbers": vin_numbers,
            }
        )

    # Workbooks for catalogs that are not in the catalog yet. Before this loop such a
    # file was dropped without a word, and a new catalog could only be added by hand
    # editing parts-data.json first.
    known = {item["catalog"] for item in old_data["catalogs"]}
    for catalog in sorted(name for name in workbook_paths if name not in known):
        metadata, uncertain = derive_catalog_metadata(catalog)
        override = overrides.get(catalog, {})
        metadata.update({key: value for key, value in override.items() if key != "catalog"})
        uncertain = [field for field in uncertain if field not in override]

        occurrences, groups, group_number, vin_numbers = parse_workbook(
            workbook_paths[catalog], metadata, group_number, fallback_figures
        )
        for item in occurrences:
            item["vinCount"] = len(vin_numbers)
            item["representativeVin"] = vin_numbers[0] if vin_numbers else ""

        all_occurrences.extend(occurrences)
        all_groups.update(groups)
        catalogs.append(
            {
                "model": metadata.get("model", ""),
                "year": metadata.get("year", ""),
                "engine": metadata.get("engine", ""),
                "vehicleType": metadata.get("vehicleType", ""),
                "catalog": catalog,
                "parts": len(occurrences),
                "vins": len(vin_numbers),
                "vinNumbers": vin_numbers,
            }
        )
        added.append(
            {
                "catalog": catalog,
                "parts": len(occurrences),
                "vins": len(vin_numbers),
                "derived": {key: metadata[key] for key in ("model", "year", "engine", "vehicleType")},
                "review": uncertain,
                "source": str(workbook_paths[catalog]),
            }
        )

    repair_blank_groups(all_groups, all_occurrences, old_data)
    parts = aggregate_parts(all_occurrences)
    linked_figures = {group["figure"] for group in all_groups.values() if group.get("figure")}
    output = {
        "generated": args.update_date,
        "catalogCount": len(catalogs),
        "figureCount": len(linked_figures),
        "occurrenceCount": len(all_occurrences),
        "uniqueParts": len(parts),
        "catalogs": catalogs,
        "groups": all_groups,
        "parts": parts,
    }
    if args.dry_run:
        print("DRY RUN - nothing written\n")
    else:
        args.output.write_text(
            json.dumps(output, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
        )

    new_part_counts = {item["catalog"]: item["parts"] for item in catalogs}
    print(f"catalogs {len(old_data['catalogs'])} -> {len(catalogs)}")
    if added:
        print(f"\nadded ({len(added)}):")
        for item in added:
            print(f"  + {item['catalog']}")
            print(f"      {item['parts']} occurrences, {item['vins']} VINs, from {item['source']}")
            print(f"      {json.dumps(item['derived'], ensure_ascii=False)}")
            if item["review"]:
                print(f"      REVIEW - could not infer {', '.join(item['review'])}; "
                      f"pass --meta to set it")
            if not item["parts"]:
                print("      WARNING - no parts parsed; check that the workbook has a detail sheet")
    if updated:
        print(f"\nupdated ({len(updated)}):")
        for catalog in updated:
            before = old_part_counts.get(catalog, 0)
            after = new_part_counts.get(catalog, 0)
            change = after - before
            print(f"  ~ {catalog}: {before} -> {after} ({change:+d})")
            if not after:
                print("      WARNING - update parsed no parts; the previous data was replaced")
    if retained:
        print(f"\nretained unchanged ({len(retained)}) - no workbook supplied")

    print(
        "\n"
        + json.dumps(
            {
                "catalogs": len(catalogs),
                "groups": len(all_groups),
                "figures": len(linked_figures),
                "occurrences": len(all_occurrences),
                "uniqueParts": len(parts),
                "output": None if args.dry_run else str(args.output),
            },
            ensure_ascii=False,
        )
    )
    if temp_context:
        temp_context.cleanup()


if __name__ == "__main__":
    main()
